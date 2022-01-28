import numpy as np
import cv2 as cv
import pytesseract
import openpyxl
import string
from copy import deepcopy
from os import scandir
from os.path import splitext, basename

# folder containing images to be matched and templates 
images_dir = ".\\images"
templates_dir = ".\\templates"
result_filename = "scan_result.xlsx"
accepted_format = [".jpg", ".jpeg", ".JPG", ".JPEG"]
mask_format = [".png", ".PNG"]

# show result images (once per image scanned)
verbose = True
preferred_height = 600 # in pixels, may vary depending on resolution
preferred_width = 900

# match threshold for template matching 
threshold = 0.48

# config for tesseract, psm 11: extract as much sparse text as possible
# upper part of button only contains uppercase letters
up_config = "--psm 11 -c tessedit_char_whitelist=" + string.ascii_uppercase
# bottom part may contain uppercase letter or digits
down_config = "--psm 11 -c tessedit_char_whitelist=" + string.ascii_uppercase + string.digits


# hardcoded since opencv takes rgb for color and making rainbow with rgb is very hard
# this also means verbose mode can only have 6 colors to distinguish betweeen templates
# in the result image, although the real scanning process will work fine with more
colors = [(0, 0, 255), (0, 120, 255), (0, 220, 255), (0, 255, 0), (255, 180, 0), (255, 0, 180)]

# prepare workbook for saving results
workbook = openpyxl.Workbook()
workbook.remove(workbook["Sheet"]) # remove the default sheet

# keys: path, image, grayscale, height, width, color
templates = list(dict())

# load template images
with scandir(templates_dir) as it:
	for entry in it:
		template = dict()
		template["path"] = entry.path

		# extract extention from filename
		if splitext(template["path"])[1] in accepted_format:
			# load accepted file into list of templates
			template["name"] = splitext(basename(template["path"]))[0]
			template["image"] = cv.imread(template["path"])
			template["height"], template["width"] =  template["image"].shape[:2]
			template["grayscale"] = cv.cvtColor(template["image"], cv.COLOR_BGR2GRAY)

			# create mask for template
			if "mask" not in template:
				template["mask"] = np.ones(template["image"].shape[:2], dtype=np.uint8)
				template["mask"][:, :] = 255
				margin = min(template["height"], template["width"]) // 10
				template["mask"][margin:margin*4, margin:margin*9] = 0 # mask over the upper text
				template["mask"][margin*6:margin*8, margin:margin*9]  = 0 # mask over the bottom text

			templates.append(template)
		elif splitext(template["path"])[1] in mask_format:
			# use user-created mask if it exists
			mask_image = cv.imread(template["path"], cv.IMREAD_UNCHANGED)
			# extract the 3rd column (alpha channel) from the image
			alpha = mask_image[:,:,3]
			alpha = np.where(alpha < 255, 0, alpha)
			template["mask"] = alpha
			cv.imshow("mask", alpha)
			cv.waitKey(0)
		

# assign color from colors array to template
if verbose:
	for template, color_index in zip(templates, range(0, len(colors), len(colors) // len(templates))):
		template["color"] = colors[color_index]

# load images 
with scandir(images_dir) as it:
	for entry in it:
		image = dict()
		image["path"] = entry.path

		# skip the file if it's not in the accepted format
		extension = splitext(image["path"])[1]
		if extension not in accepted_format:
			continue

		image["name"] = splitext(basename(image["path"]))[0]
		# prepare image to match template
		image["image"] = cv.imread(image["path"])
		image["height"], image["width"] = image["image"].shape[:2]
		image["grayscale"] = cv.cvtColor(image["image"], cv.COLOR_BGR2GRAY)
		image["result"] = deepcopy(image["image"])

		# create sheet using the name of the image being matched
		# takes up to 10 characters
		worksheet = workbook.create_sheet(image["name"][:15])
		title_column = 1 # 1 based index, increases by 2 for every template processed

		print("scanning " + image["name"])
		# loop through templates 
		for template in templates:
			print("template:", template["name"])
			# write template name to excel file
			worksheet.cell(row=1, column=title_column).value = basename(template["name"])
			entry_row = 2 # starts from the 2nd row since the first row is the title

			# get a list of locations where the image best matches the template
			result = cv.matchTemplate(image["grayscale"], template["grayscale"], cv.TM_CCOEFF_NORMED, mask=template["mask"])

			# loop through the highest matching results
			max_val = 99999 # first loop
			while max_val > threshold:
				# get the next entry
				_, max_val, _, max_loc = cv.minMaxLoc(result)

				# Prevent start_row, end_row, start_col, end_col be out of range of image
				start_row = max_loc[1] - template["height"] // 2 if max_loc[1] - template["height"] // 2 >= 0 else 0
				end_row = max_loc[1] + template["height"] // 2 + 1 if max_loc[1] + template["height"] // 2 + 1 <= image["height"] else image["height"]
				start_col = max_loc[0] - template["width"] // 2 if max_loc[0] - template["width"] // 2 >= 0 else 0
				end_col = max_loc[0] + template["width"] // 2 + 1 if max_loc[0] + template["width"] // 2 + 1 <= image["width"] else image["width"]
				
				# clear this entry from result
				result[start_row: end_row, start_col: end_col] = 0

				if max_val != np.inf and max_val > threshold:
					# draw rectangle on image to show location of button
					if verbose:
						image["result"] = cv.rectangle(image["result"],(max_loc[0],max_loc[1]), (max_loc[0]+template["width"]+1, max_loc[1]+template["height"]+1), template["color"], 2)


					# prepare image for ocr
					# crop button into up and down parts
					up_button = image["image"][ max_loc[1]:max_loc[1]+template["height"] // 2, max_loc[0]:max_loc[0]+template["width"] ].copy()
					down_button = image["image"][ max_loc[1]+template["height"] // 2:max_loc[1]+template["height"], max_loc[0]:max_loc[0]+template["width"] ].copy()
					
					# pad the image for better recognition
					margin = template["height"] // 4
					up_button = cv.copyMakeBorder(up_button, margin, margin, margin, margin, cv.BORDER_CONSTANT, None, (255, 255, 255))
					down_button = cv.copyMakeBorder(down_button, margin, margin, margin, margin, cv.BORDER_CONSTANT, None, (255, 255, 255))
					
					# dilate to clean button outline
					kernel = np.ones((2, 2), np.uint8)
					up_button = cv.dilate(up_button, kernel)
					down_button = cv.dilate(down_button, kernel)

					# cv stores image in bgr format, tesseract uses rgb
					up_button = cv.cvtColor(up_button, cv.COLOR_BGR2RGB) 					
					down_button = cv.cvtColor(down_button, cv.COLOR_BGR2RGB)

					# use tesseract to extract text, and clear whitespace from text
					up_button_text = "".join(pytesseract.image_to_string(up_button, config=up_config).split())
					down_button_text = "".join(pytesseract.image_to_string(down_button, config=down_config).split())

					if verbose == True:				
						print("button text:", up_button_text, down_button_text)
					# write result to excel file
					worksheet.cell(row=entry_row,column=title_column).value = up_button_text
					worksheet.cell(row=entry_row, column=title_column + 1).value = down_button_text
					entry_row = entry_row + 1
			title_column = title_column + 2
		# show result in one image
		if verbose == True:
			# compute scale factor
			scale_width = preferred_width / image["width"]
			scale_height = preferred_height / image["height"]
			scale = min(scale_height, scale_width)

			image["result"] = cv.resize(image["result"], (round(image["width"] * scale), round(image["height"] * scale)))
			cv.imshow("result", image["result"])
			cv.waitKey(0)
			# restore image for other templates to use
			image["result"] = deepcopy(image["image"])

print("results saved to " + result_filename)
workbook.save(result_filename)
	
